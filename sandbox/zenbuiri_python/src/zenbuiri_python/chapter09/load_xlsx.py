from openpyxl import load_workbook

target_file = 'sample/chapter09/various_worksheets.xlsx'
wb = load_workbook(target_file, read_only=True)
print(f'{target_file} is loading')

for ws in wb.worksheets:
  print(f'title: {ws.title}')
  print(f' - min_column, max_column: {ws.min_column}, {ws.max_column}')
  print(f' - min_row, max_row: {ws.min_row}, {ws.max_row}')
