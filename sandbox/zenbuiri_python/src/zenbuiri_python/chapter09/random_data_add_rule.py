import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

wb = Workbook()
ws = wb.active

values = random.sample(range(100), k=10)
for i, value in enumerate(values, 1):
  col_a = ws.cell(i, 1)
  col_a.value = value

  if i % 2 == 0:
    ws.cell(i, 2).value = i + 100

row_length = len(values)

skyblue_fill = PatternFill('solid', start_color='87CEEB', end_color='87CEEB')
orange_fill  = PatternFill('solid', start_color='FFA500', end_color='FFA500')

less_than_rule = CellIsRule(
  operator='lessThan',
  formula=[50],
  stopIfTrue=True,
  fill=skyblue_fill
)
ws.conditional_formatting.add(f'A1:A{row_length}', less_than_rule)

is_blank_rule = FormulaRule(
  formula=['ISBLANK(INDIRECT(ADDRESS(ROW(), COLUMN())))'],
  stopIfTrue=True,
  fill=orange_fill
)
ws.conditional_formatting.add(f'B1:B{row_length}', is_blank_rule)


ws.title = '条件付き書式'
wb.save('tmp/random_data_add_rule.xlsx')